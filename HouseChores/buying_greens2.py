#!/bin/python3

import openpyxl
import datetime

def main():

    list_of_occupants =['Marvinus', 'Harry', 'Mwai', 'dan']
    current_names = []
    new_list = []

    wb = openpyxl.load_workbook('/run/media/marvin/DriveEtCetera/HouseChores.xlsx')
    ws = wb.active
    current_list_for_greens(current_names,ws,wb)
    print("This is this week from monday: ",current_names)
    next_week_list(current_names,new_list,list_of_occupants)
    print(new_list)
    # writing_names_to_table(ws,new_list)
    wb.close()
    return

def current_list_for_greens(current_names,ws,wb):
    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        names = ws.cell(row=rowNumber,column=5).value
        current_names.append(names)
    return current_names

def next_week_list(current_names,new_list,list_of_occupants):
    '''Use this week's list and the member list to make new list
       list_of_occupants =['Marvinus', 'Harry', 'Mwai', 'dan']
    '''

    new_list += current_names[-4:]
    new_list.append(current_names[-4])
    return new_list

def writing_names_to_table(ws,new_list):
    'Replace column for those buying greens with new_list'
    ItemNumber = 0
    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        ws.cell(row=rowNumber,column=5).value = new_list[ItemNumber] #lynne ndero
        ItemNumber +=1

main()
