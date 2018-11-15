#!/bin/python3
#work with scheduling house chores

import random
import openpyxl
import datetime

def main():

    print("Hello,and welcome to the mopping scheduling program...")

    Occupants = ['Marvinus','Dan','Jeff','Harry']

    wb = openpyxl.load_workbook('/run/media/marvin/DriveEtCetera/HouseChores.xlsx')
    ws = wb.active

    Mopping = []#D4 downwards
    NewMopping =[]

    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        MopperName = ws.cell(row=rowNumber,column=4).value
        Mopping.append(MopperName)

    print("Now")
    NewNames(NewMopping,Mopping,Occupants)

    WriteToFile(NewMopping,ws)
    wb.save('/run/media/marvin/DriveEtCetera/NewHouseChores.xlsx')
    wb.close()
    print("\nThe new timetable has been written")
    return

def NewNames(NewMopping,Mopping,Occupants):
    '''Makes a list of new names with the last member mopping not being the first
       in the new list
    '''
    for name in range(1,8):
        NewName = random.choice(Occupants)
        NewMopping.append(NewName)
    if Mopping[-1]== NewMopping[0]:
        NewMopping[-1]== random.choice(Occupants)
    CheckTripleleNames(Occupants,NewMopping)
    return

def WriteToFile(NewMopping,ws):
    'Replace the names in the excel file for mopping'
    ItemNumber = 0
    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        ws.cell(row=rowNumber,column=4).value = NewMopping[ItemNumber] #lynne ndero
        ItemNumber +=1
    return

def CheckTripleleNames(Occupants,NewMopping):
    'Look for names appearing thrice and try to fix'
    for name in NewMopping:
        if NewMopping.count(name) > 2:
            NewMopping[NewMopping.index(name)] = random.choice(Occupants)
            CheckTripleleNames(Occupants,NewMopping)
    return
main()
