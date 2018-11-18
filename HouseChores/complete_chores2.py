#!/bin/python3

#modification of the complete_chores program
#uses method for writing,get file content old lists,
#make new_names ,write content to file

import openpyxl
import random
import os
import datetime

class HouseChores(object):
    """docstring for HouseChores"""

    # week = 1

    def __init__(self,file,occupants):
        '''initialize the class '''

        self.file = file
        self.occupants = occupants
        self.read_file()

    def read_file(self):
        "Read the Excel spreadsheet"

        self.wb = openpyxl.load_workbook(self.file)
        self.ws = self.wb.active
        self.get_file_content()

    def get_file_content(self):
        '''Gets the current lists of respective chores from file and store them
        in variables
        '''

        self.old_mopping = []
        self.old_greens = []
        self.old_water = []

        for rowNumber in range(5,self.ws.max_row+1):#skip first five rows
            self.mopper_name = self.ws.cell(row=rowNumber,column=4).value #mopping
            self.old_mopping.append(self.mopper_name)
            self.greens_name = self.ws.cell(row=rowNumber,column=5).value #greens
            self.old_greens.append(self.greens_name)
            self.fetcher_name = self.ws.cell(row=rowNumber,column=8).value #water
            self.old_water.append(self.fetcher_name)

        self.__make_new_names()

    def __make_new_names(self):
        "Uses present timetable to make list of next weeks list"
        # private because it has buying greens logic working with


        self.new_mopping = []
        self.new_greens = []
        self.new_water = []

        for name in range(1,8):
            self.new_mopper = random.choice(self.occupants)
            self.new_mopping.append(self.new_mopper)
            self.new_fetcher = random.choice(self.occupants)
            self.new_water.append(self.new_fetcher)

        # for mopping
        if self.new_mopping[-1]== self.new_mopping[0]:
            self.new_mopping[-1]== random.choice(self.occupants)

        #for water
        if self.old_water[-1]== self.new_water[0]:
            self.new_water[-1]== random.choice(self.occupants)

        #for greens
        #careful with this one
        self.new_greens += self.old_greens[-4:]
        self.new_greens += self.old_greens[-4:-1]

        print("Removing duplicates in water list...")
        self.remove_triple_names()
        self.write_to_file()

    def remove_triple_names(self):
        '''Remove names appearing more than twice in the list replacing with
        those occuring least'''

        #implement polymorphism here
        for name in self.new_water:
            if self.new_water.count(name) > 2:
                self.new_water[self.new_water.index(name)] = random.choice(self.occupants)
                self.remove_triple_names()

        #for mopping
        # print("Removing duplicates in mopping list...")
        for name in self.new_mopping:
            if self.new_mopping.count(name) > 2:
                self.new_mopping[self.new_mopping.index(name)] = random.choice(self.occupants)
                self.remove_triple_names()


    def write_to_file(self):
        "write the new timetable to a file"

        print("Finishing up..")
        print("Writing to file...")
        #for new_mopping
        self.item_number = 0

        #add duration of timetable
        self.today = datetime.date.today()
        self.duration = datetime.timedelta(days=7)
        self.end_date = self.today+self.duration
        self.today = self.today.strftime("%d-%b-%Y")
        self.end_date = self.end_date.strftime("%d-%b-%Y")
        self.ws.cell(row = 2,column=1).value = 'Dated ' + str(self.today) +" to "+str(self.end_date)

        for rowNumber in range(5,self.ws.max_row+1):#skip first five rows
            self.ws.cell(row=rowNumber,column=4).value = self.new_mopping[self.item_number] # column 4 mopping
            self.ws.cell(row=rowNumber,column=5).value = self.new_greens[self.item_number] #greens
            self.ws.cell(row=rowNumber,column=8).value = self.new_water[self.item_number] # water
            self.item_number +=1

        # HouseChores.week+=1

        # print(str(HouseChores.week)+self.file)

        self.wb.save('2'+self.file)
        print("File has been saved...")
        self.wb.close()

# main
def main():

    # file = input("Enter the path of the file: ")
    file = "HouseChores.xlsx"
    occupants =['Marvinus', 'Harry', 'Mwai', 'dan']
    timetable = HouseChores(file,occupants)

    return
main()
