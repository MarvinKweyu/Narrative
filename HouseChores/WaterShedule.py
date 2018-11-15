#!/bin/python3
#work with scheduling house chores

import random
import openpyxl
import datetime

def main():

    print("Hello,and welcome to the water scheduling program...")

    Occupants = ['Marvinus','Dan','Jeff','Harry']

    wb = openpyxl.load_workbook('/run/media/marvin/DriveEtCetera/HouseChores.xlsx')
    ws = wb.active

    BuyingGreens =[] #E4 downwards
    Water = []#H4 downwards
    NewWater = []

    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        MopperName = ws.cell(row=rowNumber,column=8).value
        Water.append(MopperName)

    print("The old list...")
    print(Water)
    NewNames(NewWater,Water,Occupants)

    print("\nThe new list...")
    print(NewWater)
    WriteToFile(NewWater,ws)
    wb.save('/run/media/marvin/DriveEtCetera/NewHouseChores.xlsx')
    wb.close()
    print("\nThe new water timetable has been written")
    return

def NewNames(NewWater,Water,Occupants):
    '''Makes a list of new names with the last member Water not being the first
       in the new list
    '''
    for name in range(1,8):
        NewName = random.choice(Occupants)
        NewWater.append(NewName)
    if Water[-1]== NewWater[0]:
        NewWater[-1]== random.choice(Occupants)
    CheckTripleleNames(Occupants,NewWater)
    return

def WriteToFile(NewWater,ws):
    'Replace the names in the excel file for Water'
    ItemNumber = 0
    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        ws.cell(row=rowNumber,column=8).value = NewWater[ItemNumber] #lynne ndero
        ItemNumber +=1
    return

def CheckTripleleNames(Occupants,NewWater):
    'Look for names appearing thrice and try to fix'
    for name in NewWater:
        if NewWater.count(name) > 2:
            NewWater[NewWater.index(name)] = random.choice(Occupants)
            CheckTripleleNames(Occupants,NewWater)
        else:
            break
    return
main()
