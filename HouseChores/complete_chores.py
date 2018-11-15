#!/bin/python3

import openpyxl
import random
import os

class HouseChores(object):
    "This is the mother class for all chores related to the house"

    def __init__(self,file,occupants):
        'initialize the class '

        self.file = file
        self.occupants = occupants
        self.read_file()

    def read_file(self):
        "Read the Excel spreadsheet"

        self.wb = openpyxl.load_workbook(self.file)
        self.ws = self.wb.active
        self.new_mopping()
        self.buying_greens()
        self.fetch_water()
        self.write_to_file()

    def new_mopping(self):
        "Get current list of those new_mopping and create the next list"

        self.new_mopping = []#D4 downwards
        print("Working on mopping...")
        
        for rowNumber in range(4,self.ws.max_row+1):#skip first four rows
            self.mopper_name = self.ws.cell(row=rowNumber,column=4).value
            self.new_mopping.append(self.mopper_name)

        #make list of new names with the last member not being the first next
        for name in range(1,8):
            self.new_name = random.choice(self.occupants)
            self.new_mopping.append(self.new_name)
        if self.new_mopping[-1]== self.new_mopping[0]:
            self.new_mopping[-1]== random.choice(self.occupants)

        self.remove_triple_names(self.new_mopping)


    def remove_triple_names(self,my_list):
        'look for names appearing thrice and fix'
        #implement polymorphism here
        print("Removing duplicates...")
        for name in my_list:
            if my_list.count(name) > 2:
                my_list[my_list.index(name)] = random.choice(self.occupants)
                # self.remove_triple_names(my_list)

    def buying_greens(self):
        "Get todays list of those buying greens ,make a new one "

        print("Working on buying greens...")
        self.current_names = []
        self.new_greens = []

        for rowNumber in range(4,self.ws.max_row+1):#skip first four rows
            self.names = self.ws.cell(row=rowNumber,column=5).value
            self.current_names.append(self.names)

        self.new_greens += self.current_names[-4:]
        # self.new_greens.append(self.current_names[-4])
        self.new_greens += self.current_names[-4:-1]

    def fetch_water(self):
        '''get current list of those fetching water,make a new one checking
        for triplets'''

        print("Working on fetching water...")
        self.water = []
        self.new_water = []

        #old list
        for rowNumber in range(4,self.ws.max_row+1):#skip first four rows
            self.fetcher_name = self.ws.cell(row=rowNumber,column=8).value
            self.water.append(self.fetcher_name)

        # new names
        for name in range(1,8):
            self.new_name = random.choice(self.occupants)
            self.new_water.append(self.new_name)
        if self.water[-1]== self.new_water[0]:
            self.new_water[-1]== random.choice(self.occupants)
        self.remove_triple_names(self.new_water)


    def write_to_file(self):
        "write the new timetable to a file"

        print("Finishing up..")
        print("Writing to file...")
        #for new_mopping
        self.item_number = 0
        print("\nThose mopping are:\n ",self.new_mopping)
        print("Those buying greens are: \n",self.new_greens)
        print("Those fetching water are: \n",self.new_water)

        for rowNumber in range(4,self.ws.max_row+1):#skip first four rows
            self.ws.cell(row=rowNumber,column=4).value = self.new_mopping[self.item_number] # column 4 mopping
            self.ws.cell(row=rowNumber,column=5).value = self.new_greens[self.item_number] #greens
            self.ws.cell(row=rowNumber,column=8).value = self.new_water[self.item_number] #lynne ndero water
            self.item_number +=1
        print("File has been saved..")
        # self.wb.save('5'+self.file)
        # self.wb.close()

# main
def main():

    # file = input("Enter the path of the file: ")
    file = "2NewHouseChores.xlsx"
    occupants =['Marvinus', 'Harry', 'Mwai', 'dan']
    timetable = HouseChores(file,occupants)

    return
main()
