#!/bin/python3

import openpyxl
import datetime


def main():

    list_of_days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    list_of_names = []

    wb = openpyxl.load_workbook('/run/media/marvin/DriveEtCetera/HouseChores.xlsx')
    ws = wb.active
    make_list_of_names(list_of_names,ws,wb)
    print(list_of_names)
    # dictionary for current green
    dictionary_of_vegies = dict(zip(list_of_days,list_of_names))
    new_dictionary_of_vegies(dictionary_of_vegies)
    wb.close()
    return

def make_list_of_names(list_of_names,ws,wb):
    for rowNumber in range(4,ws.max_row+1):#skip first four rows
        names = ws.cell(row=rowNumber,column=5).value
        list_of_names.append(names)
    return list_of_names

def new_dictionary_of_vegies(dictionary_of_vegies):
    return
main()
